import logging
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

# Configure logging
logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

# Define functions
def parse_date(date_str):
    """Parse date from various formats into a standard format."""
    if pd.isnull(date_str) or not isinstance(date_str, str):
        return date_str
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d/%m/%y", "%m/%d/%y", "%d-%b-%y", "%Y/%m/%d %I:%M %p"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return date_str

def compare_values(src_val, tgt_val):
    """Compare values from source and target, handling different data types and null values."""
    if pd.isnull(src_val) and pd.isnull(tgt_val):
        return False
    if (pd.isnull(src_val) and tgt_val == '') or (src_val == '' and pd.isnull(tgt_val)):
        return True
    if pd.isnull(src_val) or pd.isnull(tgt_val):
        return True
    if isinstance(src_val, str) and isinstance(tgt_val, str):
        src_val = parse_date(src_val)
        tgt_val = parse_date(tgt_val)
    return src_val != tgt_val

def compare_dataframes(df1, df2, primary_keys, columns_to_compare):
    try:
        df1.set_index(primary_keys, inplace=True)
        df2.set_index(primary_keys, inplace=True)

        matching_records = df1.index.intersection(df2.index)
        non_matching_records = df1.index.symmetric_difference(df2.index)

        differences = []
        for idx in matching_records:
            row1 = df1.loc[idx, columns_to_compare].reindex(columns_to_compare)
            row2 = df2.loc[idx, columns_to_compare].reindex(columns_to_compare)
            diff_cols = row1 != row2
            diff = {'key': f"{'_'.join(map(str, idx))}"}
            for col in columns_to_compare:
                if compare_values(row1[col], row2[col]):
                    diff[f'src_{col}'] = row1[col]
                    diff[f'tgt_{col}'] = row2[col]
            if len(diff) > 1:
                differences.append(diff)

        extra_records_source = df1.loc[df1.index.difference(df2.index)]
        extra_records_target = df2.loc[df2.index.difference(df1.index)]

        extra_cols_source = set(df1.columns) - set(df2.columns)
        extra_cols_target = set(df2.columns) - set(df1.columns)

        logger.info(f"Number of matching records: {len(matching_records)}")
        logger.info(f"Number of non-matching records: {len(non_matching_records)}")
        logger.info(f"Number of differences in matching records: {len(differences)}")

        return differences, extra_records_source, extra_records_target, extra_cols_source, extra_cols_target
    except Exception as e:
        logger.error(f"Error comparing dataframes: {str(e)}")
        return None, None, None, None, None

def write_output_excel(differences, extra_records_source, extra_records_target, extra_cols_source, extra_cols_target, columns_to_compare, output_file):
    wb = Workbook()
    
    # Summary of Differences sheet
    ws_diff = wb.active
    ws_diff.title = "Summary of Differences"
    ws_diff.append(['key'] + [f'src_{col}' for col in columns_to_compare] + [f'tgt_{col}' for col in columns_to_compare])
    amber_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for diff in differences:
        key = diff['key']
        row = [key]
        for col in columns_to_compare:
            row.append(diff.get(f'src_{col}', None))
        for col in columns_to_compare:
            row.append(diff.get(f'tgt_{col}', None))
        ws_diff.append(row)
        for i, cell in enumerate(ws_diff[ws_diff.max_row]):
            if i > 0 and ((pd.isna(row[i]) and pd.notna(row[i+len(columns_to_compare)])) or (pd.notna(row[i]) and pd.isna(row[i+len(columns_to_compare)]))):
                cell.fill = amber_fill
                ws_diff.cell(row=ws_diff.max_row, column=i+len(columns_to_compare)+1).fill = amber_fill
    
    # Extra Columns sheet
    ws_cols = wb.create_sheet("Extra Columns")
    ws_cols.append(["Source Extra Columns"])
    for col in extra_cols_source:
        ws_cols.append([col])
    ws_cols.append([])
    ws_cols.append(["Target Extra Columns"])
    for col in extra_cols_target:
        ws_cols.append([col])

    # Extra Records sheet
    ws_records = wb.create_sheet("Extra Records")
    ws_records.append(["Source Extra Records"])
    for _, row in extra_records_source.iterrows():
        ws_records.append(row.tolist())
    ws_records.append([])
    ws_records.append(["Target Extra Records"])
    for _, row in extra_records_target.iterrows():
        ws_records.append(row.tolist())
    
    wb.save(output_file)
    logger.info(f"Output Excel file saved: {output_file}")

def read_excel_file(file_path):
    try:
        df = pd.read_excel(file_path, parse_dates=True)
        logger.info(f"Successfully read file: {file_path}")
        return df
    except Exception as e:
        logger.error(f"Error reading file {file_path}: {str(e)}")
        return None

def main(src_file_path, tgt_file_path, primary_keys, columns_to_compare):
    output_file = f"comparison_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    df1 = read_excel_file(src_file_path)
    df2 = read_excel_file(tgt_file_path)

    if df1 is not None and df2 is not None:
        differences, extra_records_source, extra_records_target, extra_cols_source, extra_cols_target = compare_dataframes(df1, df2, primary_keys, columns_to_compare)
        if differences is not None:
            write_output_excel(differences, extra_records_source, extra_records_target, extra_cols_source, extra_cols_target, columns_to_compare, output_file)
            print(f"Comparison complete. Results saved in {output_file}")
        else:
            print("Error occurred during comparison.")
    else:
        print("Error occurred while reading files.")

if __name__ == "__main__":
    # Define the paths and parameters
    src_file_path = '/path/to/source.xlsx'
    tgt_file_path = '/path/to/target.xlsx'
    primary_keys = ['empid']  # List of primary key columns
    columns_to_compare = ['sty', 'value1', 'value2', 'dob']  # List of columns to compare
    
    # Run the main function with the provided parameters
    main(src_file_path, tgt_file_path, primary_keys, columns_to_compare)
