from lxml import etree
import openpyxl
from openpyxl.styles import PatternFill
import os
import pandas as pd

def parse_xml_to_excel(xml_file, tag_names):
    try:
        parser = etree.XMLParser(ns_clean=True, recover=True)
        tree = etree.parse(xml_file, parser)
    except etree.XMLSyntaxError as e:
        print(f"Failed to parse XML: {e}")
        return

    root = tree.getroot()
    namespaces = root.nsmap if root.nsmap else {}
    print(f"Root tag: {root.tag}")
    print(f"Namespaces: {namespaces}")

    def strip_prefix(tag):
        return tag.split('}', 1)[-1] if '}' in tag else tag.split(':')[-1]

    for tag_name in tag_names:
        if ':' in tag_name and namespaces:
            prefix, local_tag = tag_name.split(':')
            namespace_uri = namespaces.get(prefix)
            if not namespace_uri:
                print(f"Namespace for prefix '{prefix}' not found.")
                continue
            elements = root.findall(f".//{{{namespace_uri}}}{local_tag}")
        else:
            stripped_tag_name = strip_prefix(tag_name)
            elements = [elem for elem in root.iter() if strip_prefix(elem.tag) == stripped_tag_name]

        print(f"Found {len(elements)} elements for tag: {tag_name}")

        if not elements:
            print(f"No elements found for tag: {tag_name}")
            continue

        for element in elements:
            if ':' in tag_name and namespaces:
                prefix, _ = tag_name.split(':')
                namespace_uri = namespaces.get(prefix)
                obs_elements = element.findall(f".//{{{namespace_uri}}}obs")
            else:
                obs_elements = [elem for elem in element.iter() if strip_prefix(elem.tag) == 'obs']

            print(f"Found {len(obs_elements)} obs elements in {tag_name}")

            if not obs_elements:
                print(f"No obs elements found in {tag_name}")
                continue

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = strip_prefix(tag_name)

            headers = set()
            rows = []

            for obs in obs_elements:
                row = {}
                for attr, value in obs.attrib.items():
                    headers.add(attr)
                    if value == 'null':
                        row[attr] = None
                    else:
                        try:
                            row[attr] = int(value)
                        except ValueError:
                            try:
                                row[attr] = float(value)
                            except ValueError:
                                row[attr] = value

                rows.append(row)

            headers = list(headers)
            ws.append(headers)

            for row in rows:
                row_data = [row.get(header, '') for header in headers]
                ws.append(row_data)

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    if cell.value is None:
                        cell.fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

            file_name = f'/mnt/data/output_report_{strip_prefix(tag_name)}.xlsx'
            if os.path.exists(file_name):
                print(f"File {file_name} already exists and will be overwritten.")
            try:
                wb.save(file_name)
                print(f"Excel file saved: {file_name}")
            except Exception as e:
                print(f"Failed to save Excel file for {strip_prefix(tag_name)}: {e}")

